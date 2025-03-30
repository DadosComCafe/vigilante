import pandas as pd
from openpyxl import load_workbook



def get_dataframe(path: str, sep: str=",") -> pd.DataFrame|str: 
    if path.endswith(".csv"):
        return pd.read_csv(path, sep=sep)
    if path.endswith(".xlsx"):
        return pd.read_excel(path, sep=sep)
    return "Path inválido!"


def get_numeric_fields(df: pd.DataFrame) -> pd.DataFrame:
    keys = df.keys()
    dict_types = {key: str(df.dtypes[key]) for key in keys}
    list_numeric_fields = [key for key, value in dict_types.items() if value in ["int64", "float64"]]
    return df[list_numeric_fields]


def change_number_to_char(n):
    n += 1
    if 1 <= n <= 26:
        return chr(64 + n)
    else:
        return "Número fora do intervalo alfabético (1 a 26)"



def generate_report(path: str, sep: str=","):
    """Gera um arquivo xlsx com duas planilhas, uma com o conteúdo numérico do csv, e a outra a análise quantitativa"""
    #import ipdb
    #ipdb.set_trace()
    #TODO Pensar a respeito em como terminar isso
    df_data = get_dataframe(path=path, sep=sep)
    keys = df_data.keys()
    key_type_dict = {key: str(df_data.dtypes[key]) for key in keys if df_data.dtypes[key] in ('int64', 'float64')}
    list_positions = [(n, key) for n, key in enumerate(key_type_dict.keys())]
    list_positions_char = [(change_number_to_char(n), key) for n, key in list_positions]
    list_fields = [key for n, key in list_positions]
    
    file_excel = f"{path.replace('.csv', '.xlsx')}"
    sheet_name="Base"
    with pd.ExcelWriter(file_excel) as file:
        df_numeric = df_data[list_fields]
        df_numeric.to_excel(file, sheet_name=sheet_name, index=False)
        #TODO: terminar de fazer todo o cálculo com funções excel, utilizando a posição dos campos

    wb = load_workbook(file_excel)
    if "AnaliseQuantitativa" not in wb.sheetnames:
        wb.create_sheet("AnaliseQuantitativa")
    ws_calculated = wb["AnaliseQuantitativa"]
    ws_calculated["A1"] = f"=SOMA({sheet_name}.A2:A3)"
    #TODO: continuar esta ideia
    wb.save(file_excel)
