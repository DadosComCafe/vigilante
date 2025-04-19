from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
import ipdb

original_file = "assets/sample.xlsx"
numeric_new_file = original_file.replace('.xlsx', '_numeric.xlsx')

wb_original = load_workbook(original_file)

# Cria uma cópia do arquivo original
wb_original.save(numeric_new_file)
wb_new = load_workbook(numeric_new_file)

sheets = wb_original.sheetnames

for sheet_name in sheets:
    numeric_sheet_name = f"numeric_{sheet_name}"
    if numeric_sheet_name not in wb_new.sheetnames:
        wb_new.create_sheet(numeric_sheet_name)
    
    old_sheet = wb_original[sheet_name]
    numeric_sheet = wb_new[numeric_sheet_name]

    for col in old_sheet.iter_cols():
        # checando se toda a coluna é numérica
        if all(isinstance(cell.value, (int, float)) or cell.row == 1 for cell in col):
            for cell in col:
                numeric_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

# Salva o arquivo após todas as alterações
wb_new.save(numeric_new_file)

#removendo colunas vazias
for col in reversed(list(numeric_sheet.iter_cols(min_row=1, max_row=numeric_sheet.max_row, min_col=1, max_col=numeric_sheet.max_column))):
    if all(cell.value is None for cell in col):
        numeric_sheet.delete_cols(col[0].column)

wb_new.save(numeric_new_file)
#Fimo remoção de colunas

# file_final= numeric_new_file.replace("_numeric.xlsx", "_analise_quantitativa.xlsx")

# #realizando as métricas
# wb_new = load_workbook()
# #cria uma cópia
# wb_new.save(numeric_new_file)

# sheets: list = wb_new.sheetnames

# #ipdb.set_trace()
# for sheet in sheets:
#     if "numeric_" in sheet:
#         col_names = [cel.value for cel in wb_new[sheet][1]]
#         new_sheet = sheet.replace("numeric_", "AnaliseNumerica_")

#         if new_sheet not in wb_new.sheetnames:
#             wb_new.create(new_sheet)
        
#         wb_analise = wb_new[new_sheet]

#         #ipdb.set_trace()
#         #SOMATÓRIA
#         for col_index, name in enumerate(col_names, start=1):
#             col_letter = get_column_letter(col_index)
#             wb_analise[f"{col_letter}1"] = f"Somatória: {name}"
#             wb_analise[f"{col_letter}2"] = f"=SOMA({sheet}!{col_letter}:{col_letter})"

# wb_new.save(file_final)